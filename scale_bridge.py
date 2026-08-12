# -*- coding: utf-8 -*-
"""
成立规模邮箱桥接服务（龙腾鑫享产品管理平台用）
================================================
页面「📦 录入成立规模 → 📬 从邮箱读取」通过本服务读取 Coremail 邮箱中的
「基金成立汇总表」Excel，自动解析各渠道确认份额。

启动：python scale_bridge.py   （监听 127.0.0.1:8711）
依赖：openpyxl（解析 Excel）

API：
  GET /health            → {"ok": true, "user": "yangzy@chinaamc.com"}
  GET /scale?kw=576      → 找到含「龙腾鑫享576号」+「基金成立汇总表」附件的最新邮件，
                           解析返回 {"ok":true,"fund","confirmDate","total","channels":[{name,personal,inst,total}]}
  GET /find?kw=576       → 返回匹配邮件列表（不含内容）

说明：本服务只读邮箱，不修改任何邮件。授权码从 hermes secrets 文件读取，不硬编码。
"""
import imaplib, email, json, io, re, sys
from email.header import decode_header
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import urlparse, parse_qs

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SECRETS = r"D:\Users\yangzy\AppData\Local\hermes\secrets\coremail-yangzy.txt"
PORT = 8711


def load_auth():
    lines = open(SECRETS, encoding="utf-8").read().strip().splitlines()
    user = lines[0].strip()
    d = {"IMAP_USER": user}
    for line in lines[1:]:
        if "=" in line:
            k, v = line.split("=", 1)
            d[k.strip()] = v.strip()
    d.setdefault("IMAP_SERVER", "cmail.chinaamc.com:993")
    return d


def dec(s):
    if not s:
        return ""
    out = []
    for data, enc in decode_header(s):
        if isinstance(data, bytes):
            try:
                out.append(data.decode(enc or "utf-8", errors="replace"))
            except Exception:
                out.append(data.decode("utf-8", errors="replace"))
        else:
            out.append(data)
    return "".join(out)


def _num(v):
    try:
        f = float(v)
        return f if f == f else 0.0
    except (TypeError, ValueError):
        return 0.0


def parse_summary(payload):
    """解析 基金成立汇总表 xlsx → (meta, channels)"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(payload), data_only=True)
    meta = {"fund": "", "confirm_date": ""}
    channels = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        if not meta["fund"] or not meta["confirm_date"]:
            for row in rows[:6]:
                for j, c in enumerate(row):
                    if isinstance(c, str) and "基金名称" in c and j + 1 < len(row) and row[j + 1]:
                        meta["fund"] = str(row[j + 1]).strip()
                    if isinstance(c, str) and "确认日期" in c and j + 1 < len(row) and row[j + 1]:
                        meta["confirm_date"] = str(row[j + 1]).strip()
        header_idx = None
        for i, row in enumerate(rows):
            if any(isinstance(c, str) and "销售商" in c for c in row if c is not None):
                header_idx = i
                break
        if header_idx is None:
            continue
        hrow = rows[header_idx]
        name_col = next((j for j, c in enumerate(hrow) if isinstance(c, str) and "销售商" in c), 0)
        share_col = next((j for j, c in enumerate(hrow) if isinstance(c, str) and "确认份额" in c), None)
        if share_col is None:
            continue
        sub = rows[header_idx + 1] if header_idx + 1 < len(rows) else ()
        total_col = None
        for off in range(3):
            j = share_col + off
            if j < len(sub) and isinstance(sub[j], str) and sub[j].strip() == "合计":
                total_col = j
                break
        if total_col is None:
            total_col = share_col + 2
        inst_col, pers_col = total_col - 1, total_col - 2
        for row in rows[header_idx + 2:]:
            if name_col >= len(row):
                continue
            name = row[name_col]
            if name is None:
                continue
            name = str(name).strip()
            if not name or "合计" in name or "户数" in name or "制表" in name:
                continue
            total = _num(row[total_col]) if total_col < len(row) else 0.0
            if total > 0:
                channels.append({
                    "name": name,
                    "personal": _num(row[pers_col]) if pers_col < len(row) else 0.0,
                    "inst": _num(row[inst_col]) if inst_col < len(row) else 0.0,
                    "total": total,
                })
    return meta, channels


def extract_fund_no(name):
    """华夏资本龙腾鑫享576号 → 576"""
    m = re.search(r"龙腾鑫享\s*(\d+)\s*号", name or "")
    return m.group(1) if m else ""


def parse_bs_names(bs):
    """解析 BODYSTRUCTURE 字符串中的附件名（RFC2047 编码段 + name= 参数）"""
    if isinstance(bs, bytes):
        bs = bs.decode("utf-8", errors="replace")
    names = []
    # RFC2047 编码段：=?GBK?B?...?=
    for m in re.finditer(r"=\?[^?]+\?[BbQq]\?[^?]*\?=", bs):
        names.append(dec(m.group(0)))
    # name= / name*= 参数
    for m in re.finditer(r"name(\*(\d+))?\*?=([^()\s;]+)", bs):
        val = m.group(3).strip().strip('"').strip("'")
        if val.lower().startswith("utf-8''"):
            try:
                from urllib.parse import unquote
                val = unquote(val.split("''", 1)[-1])
            except Exception:
                pass
        names.append(val)
    return names


def find_mails(kw, max_scan=120):
    """扫描最近 max_scan 封主题含龙腾鑫享的邮件，返回含基金成立汇总表附件的 (uid, subject, date, fname, payload)
    策略：① 主题预筛（成立/验资/汇总/报表）② 候选邮件完整下载 ③ 附件名匹配"""
    auth = load_auth()
    server, port = auth["IMAP_SERVER"].split(":")
    M = imaplib.IMAP4_SSL(server, int(port))
    M.login(auth["IMAP_USER"], auth["IMAP_AUTH_CODE"])
    M.select("INBOX")
    hits = []
    try:
        typ, data = M.search("UTF-8", 'SUBJECT "龙腾鑫享"'.encode("utf-8"))
        if typ != "OK" or not data or not data[0]:
            return hits
        uids = [int(x) for x in data[0].split()][-max_scan:]
        # 第一轮：主题预筛（只下载 header，快）
        candidates = []
        for u in uids:
            try:
                typ, md = M.fetch(str(u), "(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE)])")
                if typ != "OK":
                    continue
                msg = email.message_from_bytes(md[0][1])
                subj = dec(msg.get("Subject") or "")
                if any(w in subj for w in ("成立", "验资", "汇总", "报表")):
                    candidates.append((u, subj, msg.get("Date") or ""))
            except Exception:
                continue
        # 第二轮：候选完整下载
        for u, subj, date in candidates:
            try:
                typ, msgdata = M.fetch(str(u), "(BODY.PEEK[])")
                if typ != "OK":
                    continue
                msg = email.message_from_bytes(msgdata[0][1])
                for part in msg.walk():
                    fn = part.get_filename()
                    if not fn:
                        continue
                    fn = dec(fn)
                    if "基金成立汇总表" not in fn:
                        continue
                    if not fn.lower().endswith((".xlsx", ".xls")):
                        continue
                    payload = part.get_payload(decode=True)
                    if not payload:
                        continue
                    if kw and kw not in fn and kw not in subj:
                        continue
                    hits.append({"uid": u, "subject": subj, "date": date, "attachment": fn, "payload": payload})
            except Exception:
                continue
        hits.sort(key=lambda h: h["uid"])
    finally:
        try:
            M.logout()
        except Exception:
            pass
    return hits


class Handler(BaseHTTPRequestHandler):
    def _send(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "*")
        self.send_header("Access-Control-Allow-Private-Network", "true")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "*")
        self.send_header("Access-Control-Allow-Private-Network", "true")
        self.end_headers()

    def do_GET(self):
        try:
            url = urlparse(self.path)
            qs = parse_qs(url.query)
            if url.path == "/health":
                self._send({"ok": True, "user": load_auth()["IMAP_USER"]})
                return
            if url.path in ("/scale", "/find"):
                kw = (qs.get("kw") or [""])[0].strip()
                if url.path == "/find":
                    hits = find_mails(kw)
                    self._send({"ok": True, "count": len(hits),
                                "mails": [{"uid": h["uid"], "subject": h["subject"], "date": h["date"], "attachment": h["attachment"]} for h in hits]})
                    return
                # /scale：解析最新命中
                hits = find_mails(kw)
                if not hits:
                    self._send({"ok": False, "error": f"未找到含「{kw}」的基金成立汇总表邮件"})
                    return
                best = hits[-1]   # uid 最大 = 最新
                meta, channels = parse_summary(best["payload"])
                total = sum(c["total"] for c in channels)
                self._send({"ok": True, "uid": best["uid"], "attachment": best["attachment"],
                            "subject": best["subject"], "date": best["date"],
                            "fund": meta["fund"], "confirmDate": meta["confirm_date"],
                            "total": total, "channels": channels})
                return
            self._send({"ok": False, "error": "404"})
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._send({"ok": False, "error": str(e)}, 500)

    def log_message(self, fmt, *args):
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))


if __name__ == "__main__":
    from http.server import ThreadingHTTPServer
    print(f"📬 成立规模邮箱桥接服务已启动: http://127.0.0.1:{PORT}")
    print("   授权邮箱:", load_auth()["IMAP_USER"])
    print("   按 Ctrl+C 停止")
    ThreadingHTTPServer(("127.0.0.1", PORT), Handler).serve_forever()
